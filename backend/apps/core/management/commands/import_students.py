"""Talabalar kontingentini Excel fayldan import qiladi (Direction/Group/AppUser).

Guruh nomi `<YO'NALISH_KODI>-<GURUH_RAQAMI>` shaklida deb qabul qilinadi (masalan
`TPI-925`) — prefiks `Direction.name`, to'liq nom `Group.name` sifatida saqlanadi.
Mos kelmagan guruh nomlari (masalan ordinatura guruhlari — kafedra nomi bilan
yozilgan) o'tkazib yuboriladi va oxirida ro'yxat qilib ko'rsatiladi.

Talaba paroli: default holatda Talaba ID (login bilan bir xil) — faqat YANGI
yaratilayotgan foydalanuvchilar uchun; mavjud foydalanuvchining paroli
o'zgartirilmaydi.

  # dry-run, faqat birinchi 10 ta qatorni sinash
  python manage.py import_students --file "data/talablar kotingenti/1-kurs.xlsx" \\
      --level-name 1-kurs --intake-year 2025 --limit 10

  # haqiqatan yozish
  python manage.py import_students --file "data/talablar kotingenti/1-kurs.xlsx" \\
      --level-name 1-kurs --intake-year 2025 --limit 10 --apply

  # butun fayl (limit yo'q)
  python manage.py import_students --file "data/talablar kotingenti/1-kurs.xlsx" \\
      --level-name 1-kurs --intake-year 2025 --apply
"""
from __future__ import annotations

import re

import bcrypt
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import AppUser, Direction, Group, Level

# Yo'nalish kodi FAQAT katta harflar bilan yoziladi (TPI, DI, MD, ЛД...) — shu orqali
# ordinatura/magistratura guruh nomlaridan (masalan "Kardiologiya-25", kichik harfli)
# ajratiladi. Guruh nomining qolgan qismi (bo'sh joy/tire/qiya chiziq bilan davom
# etuvchi raqam, kichik guruh harfi va h.k. — masalan "P-2621 A", "DI-4224A",
# "MD-101/25") e'tiborga olinmaydi, faqat kod + undan keyin darhol keladigan raqam
# borligi tekshiriladi (fullmatch emas, match — davomi ixtiyoriy).
GROUP_RE = re.compile(r"^\s*([A-ZА-Я]{1,8})[\s\-]*[0-9]")


def _hash_pw(plain: str) -> str:
    return bcrypt.hashpw(plain.encode("utf-8"), bcrypt.gensalt(rounds=10)).decode("utf-8")

# Ustun nomlari fayldan-faylga farq qiladi — hammasi shu joyda ro'yxatga olinadi.
HEADER_ALIASES = {
    "full_name": ["to‘liq ismi", "to'liq ismi", "toliq ismi"],
    "student_id": ["talaba id"],
    "passport": ["pasport raqami"],
    "pinfl": ["jshshir-kod", "jshshir kod", "jshshir"],
    "group": ["guruh", "gurux"],
    "kurs": ["kurs"],
    "faculty": ["fakultet"],
}


def normalize_header(v) -> str:
    return str(v or "").strip().lower()


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


class Command(BaseCommand):
    help = "Talabalar kontingentini Excel fayldan import qiladi (Direction/Group/AppUser)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Excel fayl yo'li (.xlsx)")
        parser.add_argument("--sheet", default="Talabalar", help="Sheet nomi (default: Talabalar)")
        parser.add_argument(
            "--level-name",
            default="",
            help="Kurs nomi, masalan '1-kurs' — faylda 'Kurs' ustuni bo'lmasa majburiy",
        )
        parser.add_argument(
            "--intake-year",
            type=int,
            required=True,
            help="Guruh qabul qilingan o'quv yili (masalan 2025) — barcha yangi guruhlarga qo'yiladi",
        )
        parser.add_argument("--limit", type=int, default=0, help="Nechta qator (0 = hammasi)")
        parser.add_argument("--apply", action="store_true", help="Haqiqatan saqlash (default: dry-run)")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl o'rnatilmagan: pip install openpyxl")

        file_path = opts["file"]
        apply_changes = bool(opts["apply"])
        limit = int(opts["limit"])
        default_level_name = str(opts["level_name"] or "").strip()
        intake_year = int(opts["intake_year"])

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = opts["sheet"]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' topilmadi. Mavjud: {wb.sheetnames}")
        ws = wb[sheet_name]

        header_row = [normalize_header(c.value) for c in ws[1]]
        col_index: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for i, h in enumerate(header_row):
                if h in aliases:
                    col_index[field] = i
                    break

        missing_required = [f for f in ("full_name", "student_id", "group") if f not in col_index]
        if missing_required:
            raise CommandError(
                f"Majburiy ustunlar topilmadi: {missing_required}. Fayl sarlavhalari: {header_row}"
            )
        if "kurs" not in col_index and not default_level_name:
            raise CommandError("Faylda 'Kurs' ustuni yo'q — --level-name bering (masalan --level-name 1-kurs)")

        def get(row_vals: list, field: str) -> str:
            i = col_index.get(field)
            return cell_str(row_vals[i]) if i is not None and i < len(row_vals) else ""

        created_students = 0
        updated_students = 0
        created_groups = 0
        created_directions: set[str] = set()
        unmatched_group: list[tuple[int, str]] = []
        missing_fields: list[int] = []
        processed = 0

        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                if limit and processed >= limit:
                    break
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                full_name = get(row_vals, "full_name")
                student_id = get(row_vals, "student_id")
                group_raw = get(row_vals, "group")

                if not full_name or not student_id or not group_raw:
                    missing_fields.append(r)
                    continue

                m = GROUP_RE.match(group_raw)
                if not m:
                    unmatched_group.append((r, group_raw))
                    continue
                direction_code = m.group(1).upper()

                level_name = get(row_vals, "kurs") or default_level_name
                level_obj, _ = Level.objects.get_or_create(name=level_name)

                direction_obj, dir_created = Direction.objects.get_or_create(name=direction_code)
                if dir_created:
                    created_directions.add(direction_code)

                group_obj, grp_created = Group.objects.get_or_create(
                    name=group_raw,
                    level=level_obj,
                    defaults={"direction": direction_obj, "intake_year": intake_year},
                )
                if grp_created:
                    created_groups += 1
                else:
                    patch_fields = []
                    if group_obj.direction_id is None:
                        group_obj.direction = direction_obj
                        patch_fields.append("direction")
                    if group_obj.intake_year is None:
                        group_obj.intake_year = intake_year
                        patch_fields.append("intake_year")
                    if patch_fields:
                        group_obj.save(update_fields=patch_fields)

                existing = AppUser.objects.filter(pk=student_id).first()
                processed += 1
                if existing:
                    changed = []
                    if existing.name != full_name:
                        existing.name = full_name
                        changed.append("name")
                    if existing.group_id != group_obj.id:
                        existing.group = group_obj
                        changed.append("group")
                    if changed:
                        updated_students += 1
                        self.stdout.write(f"  [YANGILASH] {student_id} ({full_name}): {', '.join(changed)}")
                        if apply_changes:
                            existing.save(update_fields=changed)
                    continue

                created_students += 1
                self.stdout.write(f"  [YANGI] {student_id} ({full_name}) -> {group_raw}, parol=talaba_id")
                if apply_changes:
                    AppUser.objects.create(
                        id=student_id,
                        password=_hash_pw(student_id),
                        role="student",
                        name=full_name,
                        status="Active",
                        group=group_obj,
                        profile_image="",
                    )

            if not apply_changes:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(
            f"NATIJA: yangi_talaba={created_students}  yangilangan_talaba={updated_students}  "
            f"yangi_guruh={created_groups}  yangi_yo'nalish={sorted(created_directions)}  "
            f"jarayondan_o'tgan_qator={processed}  APPLY={apply_changes}"
        )
        if missing_fields:
            self.stdout.write(self.style.WARNING(f"Majburiy maydon yo'q qatorlar: {missing_fields[:20]}"))
        if unmatched_group:
            self.stdout.write(self.style.WARNING("Guruh nomi mos kelmadi (qator, qiymat):"))
            for r, v in unmatched_group[:20]:
                self.stdout.write(f"    {r}: {v!r}")
            if len(unmatched_group) > 20:
                self.stdout.write(f"    ... yana {len(unmatched_group) - 20} ta")
        if not apply_changes:
            self.stdout.write(self.style.WARNING("Bu DRY-RUN — saqlash uchun --apply qo'shing"))
