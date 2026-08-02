"""Excel fayllardagi "Rasm katakcha ichida" (Rich Data / Picture in Cell)

formatida saqlangan talaba rasmlarini chiqarib, AppUser.profile_image ga
o'rnatadi.

Nima uchun kerak: standart Excel o'qish vositalari (openpyxl) bu formatni
"#VALUE!" deb ko'rsatadi, lekin rasmlarning o'zi fayl ichida (xl/media/*)
butun saqlangan — faqat qator <-> rasm bog'lanishi maxsus XML zanjiri orqali
(xl/metadata.xml -> xl/richData/rdrichvalue.xml -> richValueRel.xml ->
richValueRel.xml.rels) tiklanadi.

Rasm ustuni har faylda avtomatik aniqlanadi (qaysi ustunda eng ko'p `vm=`
atributli katak bo'lsa, o'sha ustun). Talaba ID import_students.py bilan bir
xil ustun nomlari orqali topiladi.

Rasm PIL bilan JPEG'ga o'giriladi (uzun tomoni max 600px, sifat 82) —
profile_image maydoni 2MB base64 chegarasidan hech qachon oshmasligi uchun.

Faqat profile_image bo'sh bo'lgan talabalarga o'rnatiladi (mavjud rasm
ustiga yozilmaydi) — agar majburan almashtirish kerak bo'lsa --overwrite.

  python manage.py seed_student_photos --file "data/talablar kotingenti/1-kurs.xlsx" --limit 5
  python manage.py seed_student_photos --file "data/talablar kotingenti/1-kurs.xlsx" --apply
"""
from __future__ import annotations

import io
import re
import zipfile
from pathlib import PurePosixPath

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.core.models import AppUser

HEADER_ALIASES = {
    "full_name": ["to‘liq ismi", "to'liq ismi", "toliq ismi"],
    "student_id": ["talaba id"],
}

MAX_DIMENSION = 600
JPEG_QUALITY = 82


def normalize_header(v) -> str:
    return str(v or "").strip().lower()


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


class RichImageIndex:
    """Bitta xlsx fayldan (bitta sheet uchun) qator->rasm bog'lanishini o'qiydi."""

    def __init__(self, xlsx_path: str, sheet_xml_name: str = "xl/worksheets/sheet1.xml"):
        self.zf = zipfile.ZipFile(xlsx_path)
        names = set(self.zf.namelist())
        if "xl/metadata.xml" not in names or "xl/richData/rdrichvalue.xml" not in names:
            self.available = False
            return
        self.available = True

        meta = self.zf.read("xl/metadata.xml").decode("utf-8")
        self.bks = [int(x) for x in re.findall(r'<xlrd:rvb i="(\d+)"/>', meta)]

        rv_xml = self.zf.read("xl/richData/rdrichvalue.xml").decode("utf-8")
        self.all_rv = re.findall(r'<rv s="(\d+)">(.*?)</rv>', rv_xml)

        rel_xml = self.zf.read("xl/richData/richValueRel.xml").decode("utf-8")
        self.rels_list = re.findall(r'<rel r:id="(rId\d+)"/>', rel_xml)

        rels_rels = self.zf.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8")
        self.rels_map = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels_rels))

        sheet_xml = self.zf.read(sheet_xml_name).decode("utf-8")
        cells = re.findall(r'<c r="([A-Z]+)(\d+)"[^>]*\bvm="(\d+)"', sheet_xml)
        col_counts: dict[str, int] = {}
        for col, _row, _vm in cells:
            col_counts[col] = col_counts.get(col, 0) + 1
        self.image_col = max(col_counts, key=col_counts.get) if col_counts else None
        self.row_to_vm = {
            int(row): int(vm) for col, row, vm in cells if col == self.image_col
        }

    def image_bytes_for_row(self, row: int) -> tuple[bytes, str] | None:
        if not self.available or self.image_col is None:
            return None
        vm = self.row_to_vm.get(row)
        if vm is None:
            return None
        bk_idx = vm - 1
        if bk_idx < 0 or bk_idx >= len(self.bks):
            return None
        rvb_i = self.bks[bk_idx]
        if rvb_i >= len(self.all_rv):
            return None
        s, body = self.all_rv[rvb_i]
        if s != "0":
            return None  # web-image — lokal fayl yo'q
        vs = re.findall(r"<v>(\d+)</v>", body)
        if not vs:
            return None
        local_id = int(vs[0])
        if local_id >= len(self.rels_list):
            return None
        rel_id = self.rels_list[local_id]
        media_path = self.rels_map.get(rel_id)
        if not media_path:
            return None
        real_path = _normalize_zip_path(str(PurePosixPath("xl/richData") / media_path))
        try:
            return self.zf.read(real_path), PurePosixPath(media_path).suffix.lstrip(".")
        except KeyError:
            return None


def _normalize_zip_path(path: str) -> str:
    """'xl/richData/../media/imageN.png' -> 'xl/media/imageN.png'."""
    parts: list[str] = []
    for part in path.split("/"):
        if part == "..":
            if parts:
                parts.pop()
        elif part and part != ".":
            parts.append(part)
    return "/".join(parts)


def to_jpeg_data_uri(raw: bytes) -> str | None:
    try:
        from PIL import Image
    except ImportError:
        return None
    try:
        img = Image.open(io.BytesIO(raw))
        img = img.convert("RGB")
        w, h = img.size
        scale = min(1.0, MAX_DIMENSION / max(w, h))
        if scale < 1.0:
            img = img.resize((max(1, int(w * scale)), max(1, int(h * scale))))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
        import base64

        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/jpeg;base64,{b64}"
    except Exception:
        return None


class Command(BaseCommand):
    help = "Excel'dagi rich-data (Picture in Cell) talaba rasmlarini AppUser.profile_image ga o'rnatadi."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Excel fayl yo'li (.xlsx)")
        parser.add_argument("--sheet-xml", default="xl/worksheets/sheet1.xml", help="Sheet XML nomi (default: sheet1.xml)")
        parser.add_argument("--limit", type=int, default=0, help="Nechta qator (0 = hammasi)")
        parser.add_argument("--apply", action="store_true", help="Haqiqatan saqlash (default: dry-run)")
        parser.add_argument("--overwrite", action="store_true", help="Mavjud profile_image ustiga ham yozish")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl o'rnatilmagan: pip install openpyxl")
        try:
            import PIL  # noqa: F401
        except ImportError:
            raise CommandError("Pillow o'rnatilmagan: pip install Pillow")

        file_path = opts["file"]
        apply_changes = bool(opts["apply"])
        overwrite = bool(opts["overwrite"])
        limit = int(opts["limit"])

        rich = RichImageIndex(file_path, opts["sheet_xml"])
        if not rich.available:
            raise CommandError("Bu faylda rich-data (Picture in Cell) topilmadi.")
        if rich.image_col is None:
            raise CommandError("Rasm ustuni aniqlanmadi (vm= atributli katak topilmadi).")
        self.stdout.write(f"Rasm ustuni avtomatik aniqlandi: {rich.image_col}\n")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        # openpyxl sheet nomini emas, sheet-xml faylini ishlatamiz — shuning uchun
        # birinchi (yoki yagona "Talabalar") sheetni olamiz.
        ws = wb["Talabalar"] if "Talabalar" in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row = [normalize_header(c.value) for c in ws[1]]
        col_index: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for i, h in enumerate(header_row):
                if h in aliases:
                    col_index[field] = i
                    break
        missing = [f for f in ("full_name", "student_id") if f not in col_index]
        if missing:
            raise CommandError(f"Majburiy ustunlar topilmadi: {missing}. Sarlavhalar: {header_row}")

        def get(row_vals: list, field: str) -> str:
            i = col_index.get(field)
            return cell_str(row_vals[i]) if i is not None and i < len(row_vals) else ""

        set_count = 0
        skipped_has_photo = 0
        skipped_no_student = 0
        skipped_no_image = 0
        skipped_web_image = 0
        skipped_encode_failed = 0
        processed = 0

        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                if limit and processed >= limit:
                    break
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                student_id = get(row_vals, "student_id")
                full_name = get(row_vals, "full_name")
                if not student_id:
                    continue
                processed += 1

                user = AppUser.objects.filter(pk=student_id).first()
                if not user:
                    skipped_no_student += 1
                    continue
                if user.profile_image and len(user.profile_image) >= 50 and not overwrite:
                    skipped_has_photo += 1
                    continue

                result = rich.image_bytes_for_row(r)
                if result is None:
                    vm = rich.row_to_vm.get(r)
                    if vm is not None:
                        bk_idx = vm - 1
                        rvb_i = rich.bks[bk_idx] if 0 <= bk_idx < len(rich.bks) else None
                        s = rich.all_rv[rvb_i][0] if rvb_i is not None and rvb_i < len(rich.all_rv) else None
                        if s == "1":
                            skipped_web_image += 1
                            continue
                    skipped_no_image += 1
                    continue

                raw, _ext = result
                data_uri = to_jpeg_data_uri(raw)
                if not data_uri or len(data_uri) < 50:
                    skipped_encode_failed += 1
                    continue

                set_count += 1
                self.stdout.write(f"  [RASM] {student_id} ({full_name}): {len(raw)}b -> {len(data_uri)}b (base64)")
                if apply_changes:
                    user.profile_image = data_uri
                    user.save(update_fields=["profile_image"])

            if not apply_changes:
                transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(
            f"NATIJA: rasm_ornatildi={set_count}  allaqachon_bor={skipped_has_photo}  "
            f"talaba_topilmadi={skipped_no_student}  rasm_topilmadi={skipped_no_image}  "
            f"web_rasm={skipped_web_image}  kodlash_xato={skipped_encode_failed}  "
            f"jarayondan_o'tgan={processed}  APPLY={apply_changes}"
        )
        if not apply_changes:
            self.stdout.write(self.style.WARNING("Bu DRY-RUN — saqlash uchun --apply qo'shing"))
